"""Regression test for XLSX redaction via the SDK.

Confirms that ``Redactor.redact_file`` routes ``.xlsx`` through the
column-aware ``pseudonymize_xlsx_smart`` pipeline (header keyword
classification + per-cell pseudonyms) instead of the flat-text fallback,
which silently dropped most entities on real spreadsheets.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest

pytestmark = pytest.mark.slow


def _build_workbook(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Email", "Notes"])
    ws.append(["Anna Müller", "anna@example.com", "leave alone"])
    ws.append(["Ben Schulz", "ben@example.com", "also untouched"])
    wb.save(path)


def test_redact_file_xlsx_uses_smart_pipeline(tmp_path):
    from openpyxl import load_workbook

    from noirdoc.sdk import Redactor

    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _build_workbook(src)

    r = Redactor(detector="presidio", language="de")
    result = r.redact_file(src, output=dst)

    assert result.reconstructed is True
    assert result.mime_type.endswith("spreadsheetml.sheet")
    assert result.entity_count >= 4

    out = load_workbook(dst)["Sheet1"]
    assert out.cell(2, 1).value.startswith("<<PERSON")
    assert out.cell(3, 1).value.startswith("<<PERSON")
    assert out.cell(2, 2).value.startswith("<<EMAIL")
    assert out.cell(3, 2).value.startswith("<<EMAIL")
    # Non-classified column passes through untouched.
    assert out.cell(2, 3).value == "leave alone"
    assert out.cell(3, 3).value == "also untouched"


def test_reveal_file_xlsx_roundtrips(tmp_path):
    from openpyxl import load_workbook

    from noirdoc.sdk import Redactor

    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _build_workbook(src)

    r = Redactor(detector="presidio", language="de")
    r.redact_file(src, output=dst)
    revealed = r.reveal_file(dst)

    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))["Sheet1"]
    assert out.cell(2, 1).value == "Anna Müller"
    assert out.cell(2, 2).value == "anna@example.com"
    assert out.cell(3, 1).value == "Ben Schulz"
    assert out.cell(3, 2).value == "ben@example.com"
