"""Smart XLSX pseudonymization using column-type inference.

Three-tier approach:
1. **Header match** – classify columns by header keywords (instant, no NLP)
2. **Sample detection** – run NLP on first N data rows for unclassified columns
3. **Skip** – columns with no PII in header or sample are ignored entirely

Only cells in classified columns are pseudonymized, using direct
``mapper.get_or_create()`` calls instead of full NLP per cell.
"""

from __future__ import annotations

import asyncio
import io
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Protocol

import structlog

if TYPE_CHECKING:
    from noirdoc.detection.base import DetectedEntity
    from noirdoc.pseudonymization.mapper import PseudonymMapper

logger = structlog.get_logger()


class _Detector(Protocol):
    """Minimal detector interface used here: just async ``detect``.

    Accepts both :class:`~noirdoc.detection.base.BaseDetector` subclasses and
    the structurally-compatible :class:`~noirdoc.detection.ensemble.EnsembleDetector`.
    """

    async def detect(self, text: str, language: str = ...) -> list[DetectedEntity]: ...


# Header keywords → entity type (substring match on normalized header)
_HEADER_ENTITY_MAP: dict[str, str] = {
    # PERSON
    "name": "PERSON",
    "nachname": "PERSON",
    "vorname": "PERSON",
    "firstname": "PERSON",
    "lastname": "PERSON",
    "full name": "PERSON",
    "patient": "PERSON",
    "mitarbeiter": "PERSON",
    "mandant": "PERSON",
    "kunde": "PERSON",
    "klient": "PERSON",
    "bewohner": "PERSON",
    "ansprechpartner": "PERSON",
    "kontaktperson": "PERSON",
    "sachbearbeiter": "PERSON",
    "betreuer": "PERSON",
    "empfänger": "PERSON",
    "absender": "PERSON",
    # EMAIL
    "email": "EMAIL",
    "e-mail": "EMAIL",
    "mail": "EMAIL",
    # PHONE
    "telefon": "PHONE",
    "tel": "PHONE",
    "phone": "PHONE",
    "handy": "PHONE",
    "mobil": "PHONE",
    "fax": "PHONE",
    "rufnummer": "PHONE",
    "durchwahl": "PHONE",
    # LOCATION
    "adresse": "LOCATION",
    "address": "LOCATION",
    "anschrift": "LOCATION",
    "wohnort": "LOCATION",
    "straße": "LOCATION",
    "strasse": "LOCATION",
    "ort": "LOCATION",
    "stadt": "LOCATION",
    "plz": "LOCATION",
    "postleitzahl": "LOCATION",
    # DATE
    "geburtsdatum": "DATE",
    "geburtstag": "DATE",
    "datum": "DATE",
    "date": "DATE",
    "birthday": "DATE",
    "geb": "DATE",
    # IBAN
    "iban": "IBAN",
    "kontonummer": "IBAN",
    "bankverbindung": "IBAN",
    # SVNR
    "sozialversicherungsnummer": "SVNR",
    "svnr": "SVNR",
    "sv-nummer": "SVNR",
    "rentenversicherungsnummer": "SVNR",
    "versicherungsnummer": "SVNR",
    # STEUER_ID
    "steuer-id": "STEUER_ID",
    "steuerid": "STEUER_ID",
    "steueridentifikationsnummer": "STEUER_ID",
    "steuernummer": "STEUER_ID",
    "identifikationsnummer": "STEUER_ID",
    "tin": "STEUER_ID",
    "idnr": "STEUER_ID",
    # ORGANIZATION
    "firma": "ORGANIZATION",
    "unternehmen": "ORGANIZATION",
    "company": "ORGANIZATION",
    "arbeitgeber": "ORGANIZATION",
    "auftraggeber": "ORGANIZATION",
    "kanzlei": "ORGANIZATION",
}


def infer_entity_type(header_value: object) -> str | None:
    """Match a header cell value against the keyword map (substring match)."""
    if not isinstance(header_value, str) or not header_value.strip():
        return None
    header_lower = header_value.strip().lower()
    for keyword, entity_type in _HEADER_ENTITY_MAP.items():
        if keyword in header_lower:
            return entity_type
    return None


@dataclass
class XlsxResult:
    """Result of smart XLSX pseudonymization."""

    new_bytes: bytes | None = None
    entity_count: int = 0
    entity_types: dict[str, int] = field(default_factory=dict)
    column_classifications: dict[str, str] = field(default_factory=dict)


async def pseudonymize_xlsx_smart(
    data: bytes,
    detector: _Detector,
    mapper: PseudonymMapper,
    language: str = "de",
    sample_rows: int = 5,
    pseudonymize: bool = True,
) -> XlsxResult:
    """Analyse and optionally pseudonymize an XLSX file using column-type inference.

    When *pseudonymize* is ``True``, cells in classified columns are replaced
    via ``mapper.get_or_create()`` and the modified workbook is returned.
    When ``False``, cells are only counted (for detect-only / block modes).
    """
    from openpyxl import load_workbook

    from noirdoc.file_analysis.extractors._zip_safety import check_ooxml_zip_safe

    result = XlsxResult()

    try:
        check_ooxml_zip_safe(data, label="xlsx")
        wb = load_workbook(io.BytesIO(data))
    except Exception as exc:
        logger.warning("xlsx_inference.load_failed", error=str(exc))
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        # --- Tier 1: classify columns from header row ---
        col_types: dict[int, str | None] = {}
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1)))
        for cell in header_row:
            etype = infer_entity_type(cell.value)
            col_types[cell.column] = etype
            if etype:
                label = cell.value if isinstance(cell.value, str) else f"col{cell.column}"
                result.column_classifications[label] = f"{etype} (header)"

        # --- Tier 2: sample first N data rows for unclassified columns ---
        unknown_cols = {col for col, t in col_types.items() if t is None}
        if unknown_cols:
            sample_cells: list[tuple[int, str]] = []
            for row in ws.iter_rows(min_row=2, max_row=min(1 + sample_rows, ws.max_row)):
                for cell in row:
                    if (
                        cell.column in unknown_cols
                        and isinstance(cell.value, str)
                        and cell.value.strip()
                    ):
                        sample_cells.append((cell.column, cell.value))

            if sample_cells:
                sem = asyncio.Semaphore(8)

                async def _detect(text: str, sem: asyncio.Semaphore = sem) -> list[DetectedEntity]:
                    async with sem:
                        return await detector.detect(text, language)

                det_results = await asyncio.gather(*[_detect(val) for _, val in sample_cells])

                for (col_idx, _), entities in zip(sample_cells, det_results, strict=False):
                    if entities and col_types.get(col_idx) is None:
                        best = max(entities, key=lambda e: e.score)
                        col_types[col_idx] = best.entity_type
                        # Find header label for logging
                        hcell = next((c for c in header_row if c.column == col_idx), None)
                        label = (
                            hcell.value
                            if hcell and isinstance(hcell.value, str)
                            else f"col{col_idx}"
                        )
                        result.column_classifications[label] = f"{best.entity_type} (sampled)"

            # Mark remaining unknowns as skip
            for col in unknown_cols:
                if col_types[col] is None:
                    col_types[col] = "skip"

        # --- Tier 3: process all data rows for classified columns ---
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                entity_type = col_types.get(cell.column)
                if not entity_type or entity_type == "skip":
                    continue
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue
                if pseudonymize:
                    cell.value = mapper.get_or_create(cell.value, entity_type)
                result.entity_count += 1
                result.entity_types[entity_type] = result.entity_types.get(entity_type, 0) + 1

    if pseudonymize and result.entity_count > 0:
        buf = io.BytesIO()
        wb.save(buf)
        result.new_bytes = buf.getvalue()

    wb.close()

    logger.info(
        "xlsx_inference.completed",
        entity_count=result.entity_count,
        entity_types=result.entity_types,
        columns=result.column_classifications,
    )

    return result
